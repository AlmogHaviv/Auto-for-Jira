#!/usr/bin/python3.9

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
import xlsxwriter
import argparse
import sys


def main():
    user_args_list = _get_user_variables()
    path_to_data_from_jira = user_args_list.data_from_jira
    path_to_priorty_excel = user_args_list.priorty_excel
    path_to_human_resource_use = user_args_list.human_resource_use
    create_sheet_for_report(path_to_data_from_jira)
    xlsx_writer(path_to_data_from_jira, budget_data(path_to_data_from_jira))
    styling_excel(path_to_data_from_jira)
    days_utilization_report(path_to_data_from_jira)
    priorty_from_excel(path_to_priorty_excel, path_to_data_from_jira)
    monthly_resource_data(path_to_human_resource_use, path_to_data_from_jira)
    days_utilization_report_data_and_calc(path_to_data_from_jira)


def _get_user_variables():
    # define format and general description
    parser = argparse.ArgumentParser(formatter_class=argparse.ArgumentDefaultsHelpFormatter,
                                     description='Report script that creates the monthly resource division'
                                                 'It also creates mid and end month report in order to monitor the sprint progress')
    parser.add_argument('-d', '--data_from_jira', type=str, default=r'C:\Users\User\PycharmProjects\Auto01\sprint_december_3.xlsx', metavar="",
                        help='From where the script will take the jira data, It has to be one-sheet excel.')
    parser.add_argument('-hr', '--human_resource_use', type=str, default=r'C:\Users\User\PycharmProjects\Auto01\4pmo1.xlsx', metavar="",
                        help="Where the script will find the data about the amount of people and how much days they will work this month and the capacity potential."
                             " \nIt has to have two sheets by the names: 'Capacity Potential by Owner' and 'Human Resources Planning' ")
    parser.add_argument('-p', '--priorty_excel', type=str, default=r'C:\Users\User\PycharmProjects\Auto01\Sprint_December_Main.xlsx', metavar="",
                        help='The excel where the script will take the priorty data from. '
                             '\npriorty data must be in column "C" ')
    parser.add_argument('-pad', '--place_algorithms_development', type=str, default='H9', metavar="",
                        help='Enter the cell place in "Human Resources Planning" where the sum of the days available for this month by algorithms development team is.')
    parser.add_argument('-pbi', '--place_bioinformatics', type=str, default='H15', metavar="",
                        help='Enter the cell place in "Human Resources Planning" where the sum of the days available for this month by bioinformatics team is.')
    parser.add_argument('-psd', '--place_software_development', type=str, default='H20', metavar="",
                        help='Enter the cell place in "Human Resources Planning" where the sum of the days available for this month by software_development team is.')
    parser.add_argument('-par', '--place_architect', type=str, default='H22', metavar="",
                        help='Enter the cell place in "Human Resources Planning" where the sum of the days available for this month by architect team is.')
    parser.add_argument('-pds', '--place_devsys', type=str, default='H22', metavar="",
                        help='Enter the cell place in "Human Resources Planning" where the sum of the days available for this month by devsys team is.')
    args = parser.parse_args(args=None if sys.argv[1:] else ['--help'])
    return args


def create_sheet_for_report(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    wb.create_sheet("sheet1")
    ws = wb["sheet1"]
    ws.append(['Notes', 'Priority', 'Tasks client/Owner', 'Task Budget Type', 'Task Resource Type/ Exec Unit', 'Task ',
               'Status', 'Approved Effort', 'Actual Status', 'Actual Effort', 'Jira ID (Issue key)', 'Task Description',
               'Task Creator/Budget Owner', 'Budget ID'])
    wb.save(path_to_data_from_jira)


# The code uses the names that has been given as the budget index, if the symbol number changes in the company - need to be updated
def budget_naming_tool(budget_type):
    if ('22' in budget_type) or ('143' in budget_type) or ('95' in budget_type) or ('144' in budget_type) or (
            'P19' and '(LavieBio)' in budget_type) or ('999' and '(LavieBio)' in budget_type):
        return 'LavieBio'
    elif ('57' in budget_type) or ('255' in budget_type):
        return 'Biomica'
    elif ('209' in budget_type) or ('198' in budget_type) or ('197' in budget_type) or ('205' in budget_type) or (
            '199' in budget_type) or (
            '150' in budget_type):
        return 'Canonic'
    elif ('21' in budget_type) or ('211' in budget_type) or ('210' in budget_type) or ('23' in budget_type):
        return 'Agplenus'
    elif ('P64' in budget_type) or ('999' and '(AgSeeds/IC)' in budget_type) or ('157' in budget_type) or (
            '162' in budget_type) or (
            '24' in budget_type) or ('8' in budget_type) or ('129' in budget_type) or ('128' in budget_type):
        return 'Agseed'
    elif '269' in budget_type:
        return 'CSO'
    elif '265' in budget_type:
        return 'CSO Microbials'
    elif '264' in budget_type:
        return 'Chempass'
    elif ('277' in budget_type) or ('223' in budget_type) or ('266' in budget_type) or ('278' in budget_type):
        return 'Generator'
    elif '274' in budget_type:
        return 'CP Upkeep'
    elif '272' in budget_type:
        return 'GR Upkeep'
    elif '273' in budget_type:
        return 'MB Upkeep'
    elif '276' in budget_type:
        return 'CPBC'
    elif '271' in budget_type:
        return 'CPBC Upkeep'
    elif '275' in budget_type:
        return 'CPBE Upkeep'
    elif '280' in budget_type:
        return 'Generator'
    else:
        return 'Error'


# list of names for the resource types
def task_resource_tool(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    cs = wb.worksheets[0]
    raw_name_data = []
    df = pd.read_excel(path_to_data_from_jira)
    col_no = df.columns.get_loc('Assignee')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        raw_name_data.append(cell.value)
    raw_name_data.pop(0)
    wb.save(path_to_data_from_jira)
    list_of_developers_names = []
    for val in raw_name_data:
        if val is not None:
            list_of_developers_names.append(val)
        else:
            list_of_developers_names.append(0)
    return list_of_developers_names


# converting the list of names to the resource types
def task_resource_list(name_of_assignee):
    if ('robertoo' in name_of_assignee) or ('renanam' in name_of_assignee) or ('anatm' in name_of_assignee) or (
            'itair' in name_of_assignee):
        return f'Algorithms Development ({name_of_assignee})'
    elif ('gala' in name_of_assignee) or ('markb' in name_of_assignee) or ('michala' in name_of_assignee) or (
            'iliab' in name_of_assignee) or (
            'jonathans' in name_of_assignee) or ('liorr' in name_of_assignee):
        return f'Bioinformatics ({name_of_assignee})'
    elif ('duduz' in name_of_assignee) or ('noama' in name_of_assignee) or ('nerias' in name_of_assignee) or (
            'hodayam' in name_of_assignee):
        return f'Software Development ({name_of_assignee})'
    elif 'taln' in name_of_assignee:
        return f'Architect ({name_of_assignee})'
    elif ('andreyl' in name_of_assignee) or ('liavs' in name_of_assignee) or ('vladz' in name_of_assignee):
        return f'Dev/SysOps ({name_of_assignee})'
    elif 'iliaz' in name_of_assignee:
        return f'Director of CPBC ({name_of_assignee})'
    else:
        return f'other - {name_of_assignee}'


# naming the third row by task budget type
def naming_tool(name_of_client):
    if name_of_client == 'CP Upkeep':
        return 'Chempass'
    elif name_of_client == 'CPBC Upkeep':
        return 'CPBC'
    elif name_of_client == 'MB Upkeep':
        return 'CSO Microbials'
    elif name_of_client == 'CPBE Upkeep':
        return 'CPBC'
    else:
        return name_of_client


# Gets from the jira data the budget ID
def budget_data(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    cs = wb.worksheets[0]
    raw_name_data = []
    df = pd.read_excel(path_to_data_from_jira)
    col_no = df.columns.get_loc('Custom field (Budget)')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        raw_name_data.append(cell.value)
    raw_name_data.pop(0)
    wb.save(path_to_data_from_jira)
    list_of_budget_id = []
    for val in raw_name_data:
        if val is not None:
            list_of_budget_id.append(val)
        else:
            list_of_budget_id.append(0)
    return list_of_budget_id


# Gets from the jira data the time estimation for a mission
def task_effort_estimation_tool(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    cs = wb.worksheets[0]
    raw_name_data = []
    df = pd.read_excel(path_to_data_from_jira)
    col_no = df.columns.get_loc('Original Estimate')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        raw_name_data.append(cell.value)
    raw_name_data.pop(0)
    wb.save(path_to_data_from_jira)
    res = []
    for val in raw_name_data:
        try:
            new_val = int(val) / 3600
            new_val1 = new_val / 8
            res.append(round(new_val1, 2))
        except TypeError:
            res.append(0)
    return res


# Gets from the jira data the jira ID
def jira_id(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    cs = wb.worksheets[0]
    jira_id_data = []
    df = pd.read_excel(path_to_data_from_jira)
    col_no = df.columns.get_loc('Issue key')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        jira_id_data.append(cell.value)
    jira_id_data.pop(0)
    wb.save(path_to_data_from_jira)
    res_jira_id = []
    for val in jira_id_data:
        if val is not None:
            res_jira_id.append(val)
        else:
            res_jira_id.append(0)
    return res_jira_id


# not in use - need to get data from, Jira first
def priorty_func(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    cs = wb.worksheets[0]
    raw_priorty_data = []
    df = pd.read_excel(path_to_data_from_jira)
    col_no = df.columns.get_loc('Priorty')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        raw_priorty_data.append(cell.value)
    raw_priorty_data.pop(0)
    wb.save(path_to_data_from_jira)
    res_priorty = []
    for val in raw_priorty_data:
        if None != val:
            res_priorty.append(val)
        else:
            res_priorty.append(0)
    return res_priorty


# Gets from the jira data the task description
def task_description(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    cs = wb.worksheets[0]
    task_description_data = []
    df = pd.read_excel(path_to_data_from_jira)
    col_no = df.columns.get_loc('Summary')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        task_description_data.append(cell.value)
    task_description_data.pop(0)
    wb.save(path_to_data_from_jira)
    res_task_description_data = []
    for val in task_description_data:
        if None != val:
            res_task_description_data.append(val)
        else:
            res_task_description_data.append(0)
    return res_task_description_data


# Gets from the jira data the task creator
def task_creator(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    cs = wb.worksheets[0]
    task_creator_data = []
    df = pd.read_excel(path_to_data_from_jira)
    col_no = df.columns.get_loc('Creator')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        task_creator_data.append(cell.value)
    task_creator_data.pop(0)
    wb.save(path_to_data_from_jira)
    res_task_creator_data = []
    for val in task_creator_data:
        if None != val:
            res_task_creator_data.append(val)
        else:
            res_task_creator_data.append(0)
    return res_task_creator_data


# Decides if missions are approved or not and insert then into G and F rows and color the rows
def monthly_resource_func(name_of_company_resource, days_of_work, letter, worksheet):
    if name_of_company_resource >= days_of_work:
        worksheet['G' + str(letter)].value = 'approved'
        worksheet['H' + str(letter)].value = worksheet['F' + str(letter)].value
        worksheet['G' + str(letter)].fill = PatternFill(patternType='solid', fgColor='75EE15')
    elif (name_of_company_resource <= days_of_work) and (name_of_company_resource > 0):
        worksheet['G' + str(letter)].value = 'approved'
        worksheet['H' + str(letter)].value = name_of_company_resource
        worksheet['A' + str(
            letter)].value = f"The task need to be split. {round((days_of_work - name_of_company_resource), 2)} are left for next month. "
        worksheet['G' + str(letter)].fill = PatternFill(patternType='solid', fgColor='ECB918')
    else:
        worksheet['G' + str(letter)].value = 'Postponed'
        worksheet['H' + str(letter)].value = 0
        worksheet['G' + str(letter)].fill = PatternFill(patternType='solid', fgColor='FC2C03')


def days_of_work_per_company(path_to_human_resource_use, place_for_company_budget, resource_type, place_for_amount_of_days_per_team):
    workbook = load_workbook(path_to_human_resource_use, data_only=True)
    worksheet1 = workbook['Capacity Potential by Owner']
    worksheet3 = workbook['Human Resources Planning']
    algorithms_development_total_potential = worksheet1['T3'].value - worksheet1['S3'].value
    bioinformatics_total_potential = worksheet1['T4'].value - worksheet1['S4'].value
    software_development_total_potential = worksheet1['T5'].value - worksheet1['S5'].value
    architect_total_potential = worksheet1['T6'].value
    devsys_total_potential = worksheet1['T7'].value - worksheet1['S7'].value
    amount_of_days_per_team = worksheet3[place_for_amount_of_days_per_team].value
    if worksheet1[place_for_company_budget].value is None:
        company_budget = 0
    else:
        company_budget = worksheet1[place_for_company_budget].value
    if resource_type == 'algorithms_development':
        amount_of_days_per_company = round(((company_budget / algorithms_development_total_potential) * amount_of_days_per_team), 2)
    elif resource_type == 'bioinformatics':
        amount_of_days_per_company = round(((company_budget / bioinformatics_total_potential) * amount_of_days_per_team), 2)
    elif resource_type == 'software_development':
        amount_of_days_per_company = round(((company_budget / software_development_total_potential) * amount_of_days_per_team), 2)
    elif resource_type == 'architect':
        amount_of_days_per_company = round(((company_budget / architect_total_potential) * amount_of_days_per_team), 2)
    elif resource_type == 'devsys':
        amount_of_days_per_company = round(((company_budget / devsys_total_potential) * amount_of_days_per_team), 2)
    else:
        amount_of_days_per_company = 0
    return amount_of_days_per_company


# creates the data in order to decide which missions are approved and which ,missions will be declined.
# creates the second sheet in the report where we can get some overall data for annual calculations
# uses a parser in order to get locations of the totals in human resource
def monthly_resource_data(path_to_human_resource_use, path_to_data_from_jira):
    workbook_for_report = load_workbook(path_to_data_from_jira, data_only=True)
    place_of_algorithms_development_total_days = 'H9'
    place_of_bioinformatics_total_days = 'H15'
    place_of_software_development_total_days = 'H20'
    place_of_architect_total_days = 'H22'
    place_of_devsys_total_days = 'H26'
    worksheet_for_report = workbook_for_report['Sprint Planning Report']
    lavie_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'B3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['C3'] = lavie_bio_algo
    lavie_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'B4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['C11'] = lavie_bio_bioinfo
    biomica_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'C3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['D3'] = biomica_bio_algo
    biomica_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'C4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['D11'] = biomica_bio_bioinfo
    biomica_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'C5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['D7'] = biomica_bio_softwaredev
    canonic_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'D4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['E11'] = canonic_bio_bioinfo
    canonic_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'D3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['E3'] = canonic_bio_algo
    canonic_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'D5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['E7'] = canonic_bio_softwaredev
    agplenus_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'E4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['F11'] = agplenus_bio_bioinfo
    agplenus_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'E3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['F3'] = agplenus_bio_algo
    agplenus_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'E5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['F7'] = agplenus_bio_softwaredev
    agseed_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'G3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['H3'] = agseed_bio_algo
    agseed_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'G4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['H11'] = agseed_bio_bioinfo
    agseed_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'G5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['H7'] = agseed_bio_softwaredev
    cso_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'H5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['J7'] = cso_bio_softwaredev
    csomicrobials_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'I4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['k11'] = csomicrobials_bio_bioinfo
    csomicrobials_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'I3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['K3'] = csomicrobials_bio_algo
    csomicrobials_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'I5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['K7'] = csomicrobials_bio_softwaredev
    csomicrobials_bio_devsys = days_of_work_per_company(path_to_human_resource_use, 'I7', 'devsys', place_of_devsys_total_days)
    mbupkeep_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'J4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['L11'] = mbupkeep_bio_bioinfo
    mbupkeep_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'J3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['L3'] = mbupkeep_bio_algo
    mbupkeep_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'J5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['L7'] = mbupkeep_bio_softwaredev
    chempass_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'K3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['M3'] = chempass_bio_algo
    chempass_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'K5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['M7'] = chempass_bio_softwaredev
    cpupkeep_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'L4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['N11'] = cpupkeep_bio_bioinfo
    cpupkeep_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'L3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['N3'] = cpupkeep_bio_algo
    cpupkeep_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'L5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['N7'] = cpupkeep_bio_softwaredev
    cpupkeep_bio_devsys = days_of_work_per_company(path_to_human_resource_use, 'L7', 'devsys', place_of_devsys_total_days)
    genertor_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'M4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['O11'] = genertor_bio_bioinfo
    genertor_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'M3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['O3'] = genertor_bio_algo
    genertor_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'M5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['O7'] = genertor_bio_algo
    genertor_bio_devsys = days_of_work_per_company(path_to_human_resource_use, 'M7', 'devsys', place_of_devsys_total_days)
    grupkeep_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'N4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['P11'] = grupkeep_bio_bioinfo
    grupkeep_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'N3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['P3'] = grupkeep_bio_algo
    grupkeep_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'N5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['P7'] = grupkeep_bio_softwaredev
    cpbc_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'O4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['R11'] = cpbc_bio_bioinfo
    cpbc_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'O3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['R3'] = cpbc_bio_algo
    cpbc_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'O5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['R7'] = cpbc_bio_softwaredev
    cpbc_bio_devsys = days_of_work_per_company(path_to_human_resource_use, 'O7', 'devsys', place_of_devsys_total_days)
    cpbc_bio_architect = days_of_work_per_company(path_to_human_resource_use, 'O6', 'architect', place_of_architect_total_days)
    cpbcupkeep_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'P4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['S11'] = cpbcupkeep_bio_bioinfo
    cpbcupkeep_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'P3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['S3'] = cpbcupkeep_bio_algo
    cpbcupkeep_bio_architect = days_of_work_per_company(path_to_human_resource_use, 'P6', 'architect', place_of_architect_total_days)
    cpbcupkeep_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'P5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['S7'] = cpbcupkeep_bio_softwaredev
    cpbcupkeep_bio_devsys = days_of_work_per_company(path_to_human_resource_use, 'P7', 'devsys', place_of_devsys_total_days)
    cpbe_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'Q4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['T11'] = cpbe_bio_bioinfo
    cpbe_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'Q3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['T3'] = cpbe_bio_algo
    cpbe_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'Q5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['T7'] = cpbe_bio_softwaredev
    cpbe_bio_devsys = days_of_work_per_company(path_to_human_resource_use, 'Q7', 'devsys', place_of_devsys_total_days)
    cpbeupkeep_bio_bioinfo = days_of_work_per_company(path_to_human_resource_use, 'R4', 'bioinformatics', place_of_bioinformatics_total_days)
    worksheet_for_report['U11'] = cpbeupkeep_bio_bioinfo
    cpbeupkeep_bio_algo = days_of_work_per_company(path_to_human_resource_use, 'R3', 'algorithms_development', place_of_algorithms_development_total_days)
    worksheet_for_report['U3'] = cpbeupkeep_bio_algo
    cpbeupkeep_bio_softwaredev = days_of_work_per_company(path_to_human_resource_use, 'R5', 'software_development', place_of_software_development_total_days)
    worksheet_for_report['U7'] = cpbeupkeep_bio_softwaredev
    cpbeupkeep_bio_devsys = days_of_work_per_company(path_to_human_resource_use, 'R7', 'devsys', place_of_devsys_total_days)
    ws2 = workbook_for_report['Sheet1']
    full_row_data = []
    for number in range(3, ws2.max_row + 1):
        row_data = []
        for cell in ws2[number]:
            row_data.append(cell.value)
        full_row_data.append(row_data)
    # sort the data in a way that the priorty 1 is always firs to be checked in order to be approved first.
    sort_full_row_data = sorted(full_row_data, key=lambda priorty: priorty[1])
    for row_data in sort_full_row_data:
        number = row_data[14]
        if row_data[3] == 'LavieBio':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(lavie_bio_bioinfo, row_data[5], number, ws2)
                lavie_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(lavie_bio_algo, row_data[5], number, ws2)
                lavie_bio_algo -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'Biomica':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(biomica_bio_bioinfo, row_data[5], number, ws2)
                biomica_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(biomica_bio_algo, row_data[5], number, ws2)
                biomica_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(biomica_bio_softwaredev, row_data[5], number, ws2)
                biomica_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'Canonic':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(canonic_bio_bioinfo, row_data[5], number, ws2)
                canonic_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(canonic_bio_algo, row_data[5], number, ws2)
                canonic_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(canonic_bio_softwaredev, row_data[5], number, ws2)
                canonic_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'Agplenus':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(agplenus_bio_bioinfo, row_data[5], number, ws2)
                agplenus_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(agplenus_bio_algo, row_data[5], number, ws2)
                agplenus_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(agplenus_bio_softwaredev, row_data[5], number, ws2)
                agplenus_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'Agseed':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(agseed_bio_bioinfo, row_data[5], number, ws2)
                agseed_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(agseed_bio_algo, row_data[5], number, ws2)
                agseed_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(agseed_bio_softwaredev, row_data[5], number, ws2)
                agseed_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'CSO Microbials':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(csomicrobials_bio_bioinfo, row_data[5], number, ws2)
                csomicrobials_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(csomicrobials_bio_algo, row_data[5], number, ws2)
                csomicrobials_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(csomicrobials_bio_softwaredev, row_data[5], number, ws2)
                csomicrobials_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(csomicrobials_bio_devsys, row_data[5], number, ws2)
                csomicrobials_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'MB Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(mbupkeep_bio_bioinfo, row_data[5], number, ws2)
                mbupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(mbupkeep_bio_algo, row_data[5], number, ws2)
                mbupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(mbupkeep_bio_softwaredev, row_data[5], number, ws2)
                mbupkeep_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'Chempass':
            if 'Algorithms Development' in row_data[4]:
                monthly_resource_func(chempass_bio_algo, row_data[5], number, ws2)
                chempass_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(chempass_bio_softwaredev, row_data[5], number, ws2)
                chempass_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'CP Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpupkeep_bio_bioinfo, row_data[5], number, ws2)
                cpupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpupkeep_bio_algo, row_data[5], number, ws2)
                cpupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpupkeep_bio_softwaredev, row_data[5], number, ws2)
                cpupkeep_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpupkeep_bio_devsys, row_data[5], number, ws2)
                cpupkeep_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'Generator':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(genertor_bio_bioinfo, row_data[5], number, ws2)
                genertor_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(genertor_bio_algo, row_data[5], number, ws2)
                genertor_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(genertor_bio_softwaredev, row_data[5], number, ws2)
                genertor_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(genertor_bio_devsys, row_data[5], number, ws2)
                genertor_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'GR Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(grupkeep_bio_bioinfo, row_data[5], number, ws2)
                grupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(grupkeep_bio_algo, row_data[5], number, ws2)
                grupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(grupkeep_bio_softwaredev, row_data[5], number, ws2)
                grupkeep_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'CPBC':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpbc_bio_bioinfo, row_data[5], number, ws2)
                cpbc_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpbc_bio_algo, row_data[5], number, ws2)
                cpbc_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpbc_bio_softwaredev, row_data[5], number, ws2)
                cpbc_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpbc_bio_devsys, row_data[5], number, ws2)
                cpbc_bio_devsys -= row_data[5]
            elif 'Architect' in row_data[4]:
                monthly_resource_func(cpbc_bio_architect, row_data[5], number, ws2)
                cpbc_bio_architect -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'CPBC Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_bioinfo, row_data[5], number, ws2)
                cpbcupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_algo, row_data[5], number, ws2)
                cpbcupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_softwaredev, row_data[5], number, ws2)
                cpbcupkeep_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_devsys, row_data[5], number, ws2)
                cpbcupkeep_bio_devsys -= row_data[5]
            elif 'Architect' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_architect, row_data[5], number, ws2)
                cpbcupkeep_bio_architect -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'CSO':
            if 'Software Development' in row_data[4]:
                monthly_resource_func(cso_bio_softwaredev, row_data[5], number, ws2)
                cso_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'CPBE':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpbe_bio_bioinfo, row_data[5], number, ws2)
                cpbe_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpbe_bio_algo, row_data[5], number, ws2)
                cpbe_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpbe_bio_softwaredev, row_data[5], number, ws2)
                cpbe_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpbe_bio_devsys, row_data[5], number, ws2)
                cpbe_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        elif row_data[3] == 'CPBE Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpbeupkeep_bio_bioinfo, row_data[5], number, ws2)
                cpbeupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpbeupkeep_bio_algo, row_data[5], number, ws2)
                cpbeupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpbeupkeep_bio_softwaredev, row_data[5], number, ws2)
                cpbeupkeep_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpbeupkeep_bio_devsys, row_data[5], number, ws2)
                cpbeupkeep_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(number)].value = 'approved - no evogene resources'
                ws2['H' + str(number)].value = ws2['F' + str(number)].value
        else:
            ws2['G' + str(number)].value = 'approved - no evogene resources'
            ws2['H' + str(number)].value = ws2['F' + str(number)].value
    worksheet_for_report['C4'] = worksheet_for_report['C3'].value - lavie_bio_algo
    worksheet_for_report['D4'] = worksheet_for_report['D3'].value - biomica_bio_algo
    worksheet_for_report['E4'] = worksheet_for_report['E3'].value - canonic_bio_algo
    worksheet_for_report['F4'] = worksheet_for_report['F3'].value - agplenus_bio_algo
    worksheet_for_report['H4'] = worksheet_for_report['H3'].value - agseed_bio_algo
    worksheet_for_report['K4'] = worksheet_for_report['K3'].value - csomicrobials_bio_algo
    worksheet_for_report['L4'] = worksheet_for_report['L3'].value - mbupkeep_bio_algo
    worksheet_for_report['M4'] = worksheet_for_report['M3'].value - chempass_bio_algo
    worksheet_for_report['N4'] = worksheet_for_report['N3'].value - cpupkeep_bio_algo
    worksheet_for_report['O4'] = worksheet_for_report['O3'].value - genertor_bio_algo
    worksheet_for_report['P4'] = worksheet_for_report['P3'].value - grupkeep_bio_algo
    worksheet_for_report['R4'] = worksheet_for_report['R3'].value - cpbc_bio_algo
    worksheet_for_report['S4'] = worksheet_for_report['S3'].value - cpbcupkeep_bio_algo
    worksheet_for_report['T4'] = worksheet_for_report['T3'].value - cpbe_bio_algo
    worksheet_for_report['U4'] = worksheet_for_report['U3'].value - cpbeupkeep_bio_algo
    worksheet_for_report['C12'] = worksheet_for_report['C11'].value - lavie_bio_bioinfo
    worksheet_for_report['D12'] = worksheet_for_report['D11'].value - biomica_bio_bioinfo
    worksheet_for_report['D8'] = worksheet_for_report['D7'].value - biomica_bio_softwaredev
    worksheet_for_report['E12'] = worksheet_for_report['E11'].value - canonic_bio_bioinfo
    worksheet_for_report['E8'] = worksheet_for_report['E7'].value - canonic_bio_softwaredev
    worksheet_for_report['F12'] = worksheet_for_report['F11'].value - agplenus_bio_bioinfo
    worksheet_for_report['F8'] = worksheet_for_report['F7'].value - agplenus_bio_softwaredev
    worksheet_for_report['H12'] = worksheet_for_report['H11'].value - agseed_bio_bioinfo
    worksheet_for_report['H8'] = worksheet_for_report['H7'].value - agseed_bio_softwaredev
    worksheet_for_report['J8'] = worksheet_for_report['J7'].value - cso_bio_softwaredev
    worksheet_for_report['K12'] = worksheet_for_report['K11'].value - csomicrobials_bio_bioinfo
    worksheet_for_report['K8'] = worksheet_for_report['K7'].value - csomicrobials_bio_softwaredev
    worksheet_for_report['L12'] = worksheet_for_report['L11'].value - mbupkeep_bio_bioinfo
    worksheet_for_report['L8'] = worksheet_for_report['L7'].value - mbupkeep_bio_softwaredev
    worksheet_for_report['M8'] = worksheet_for_report['M7'].value - chempass_bio_softwaredev
    worksheet_for_report['N12'] = worksheet_for_report['N11'].value - cpupkeep_bio_bioinfo
    worksheet_for_report['N8'] = worksheet_for_report['N7'].value - cpupkeep_bio_softwaredev
    worksheet_for_report['O12'] = worksheet_for_report['O11'].value - genertor_bio_bioinfo
    worksheet_for_report['P12'] = worksheet_for_report['P11'].value - grupkeep_bio_bioinfo
    worksheet_for_report['P8'] = worksheet_for_report['P7'].value - grupkeep_bio_softwaredev
    worksheet_for_report['R12'] = worksheet_for_report['R11'].value - cpbc_bio_bioinfo
    worksheet_for_report['R8'] = worksheet_for_report['R7'].value - cpbc_bio_softwaredev
    worksheet_for_report['S12'] = worksheet_for_report['S11'].value - cpbcupkeep_bio_bioinfo
    worksheet_for_report['S8'] = worksheet_for_report['S7'].value - cpbcupkeep_bio_softwaredev
    worksheet_for_report['T12'] = worksheet_for_report['T11'].value - cpbe_bio_bioinfo
    worksheet_for_report['T8'] = worksheet_for_report['T7'].value - cpbe_bio_softwaredev
    worksheet_for_report['U12'] = worksheet_for_report['U11'].value - cpbeupkeep_bio_bioinfo
    worksheet_for_report['U8'] = worksheet_for_report['U7'].value - cpbeupkeep_bio_softwaredev
    workbook_for_report.save(path_to_data_from_jira)
    return sort_full_row_data


# accumulate all the data from the functions and writes the data into the Excel sheet
def xlsx_writer(path_to_data_from_jira, raw_name_data):
    jira_id_data = jira_id(path_to_data_from_jira)
    task_description_data = task_description(path_to_data_from_jira)
    task_creator_data = task_creator(path_to_data_from_jira)
    task_resource_data = []
    actual_status_data = []
    actual_effort_data = []
    excel_row_number = []
    count = 2
    for string in task_resource_tool(path_to_data_from_jira):
        new_string = str(string)
        task_resource_data.append(task_resource_list(new_string))
        actual_status_data.append('Pending/Planned')
        actual_effort_data.append('0')
    task_effort_estimation_data = task_effort_estimation_tool(path_to_data_from_jira)
    wb1 = xlsxwriter.Workbook(path_to_data_from_jira)
    worksheet = wb1.add_worksheet()
    bold = wb1.add_format({'bold': True})
    headers = ['Notes', 'Priority', 'Tasks client/Owner', 'Task Budget Type', 'Task Resource Type/ Exec Unit',
               'Task Effort Estimation (Days)',
               'Task Status', 'Approved Effort', 'Actual Status', 'Actual Effort', 'Jira ID (Issue key)',
               'Task Description',
               'Task Creator/Budget Owner', 'Budget ID', 'Excel Row']
    worksheet.write_row(1, 0, headers, bold)
    budget_name_data = []
    for item in raw_name_data:
        new_item = str(item)
        budget_name_data.append(budget_naming_tool(new_item))
        count += 1
        excel_row_number.append(count)
    name_data = []
    for path_to_data_from_jira in budget_name_data:
        name_data.append(naming_tool(path_to_data_from_jira))
    worksheet.write_column(2, 13, raw_name_data)
    worksheet.write_column(2, 3, budget_name_data)
    worksheet.write_column(2, 2, name_data)
    worksheet.write_column(2, 10, jira_id_data)
    worksheet.write_column(2, 11, task_description_data)
    worksheet.write_column(2, 12, task_creator_data)
    worksheet.write_column(2, 4, task_resource_data)
    worksheet.write_column(2, 5, task_effort_estimation_data)
    worksheet.write_column(2, 8, actual_status_data)
    worksheet.write_column(2, 9, actual_effort_data)
    worksheet.write_column(2, 14, excel_row_number)
    wb1.close()


# give the style of the Excel - it responsible for all the colores and fonts except the bold headers and the colors of approved or postponed missions.
def styling_excel(name_of_workbook):
    wb = load_workbook(name_of_workbook)
    ws = wb['Sheet1']
    for col in range(1, 16):
        ws[get_column_letter(col) + '1'].fill = PatternFill(patternType='solid', fgColor='38D9D1')
    for col in range(1, 16):
        ws[get_column_letter(col) + '2'].fill = PatternFill(patternType='solid', fgColor='DAF7A6')
    ws['C1'] = 'Planning'
    ws.merge_cells('C1:H1')
    ws['I1'] = 'Status'
    ws.merge_cells('I1:J1')
    ws['K1'] = 'ID'
    ws['L1'] = 'Description'
    ws.merge_cells('L1:O1')
    font_style = Font(name='Verdana', size=10)
    a = Side(border_style='medium', color='000404')
    border = Border(top=a, bottom=a, left=a, right=a)
    for row in ws:
        for cell in row:
            cell.font = font_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    ws['C1'] = 'Planning'
    ws.merge_cells('C1:H1')
    wb.save(name_of_workbook)


# creates the second sheet in the report excel where we are getting to see all the days used and requested by the companies.
def days_utilization_report(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    worksheet = wb.create_sheet('Sheet_B')
    worksheet.title = 'Sprint Planning Report'
    for col in range(1, 24):
        worksheet[get_column_letter(col) + '1'].fill = PatternFill(patternType='solid', fgColor='DAF7A6')
        worksheet[get_column_letter(col) + '2'].fill = PatternFill(patternType='solid', fgColor='DAF7A6')
    worksheet['A1'] = 'Resource Type'
    worksheet.merge_cells('A1:A2')
    worksheet['C1'] = 'Exogenic Customers Budget'
    worksheet.merge_cells('C1:I1')
    worksheet['J1'] = 'Product Budget'
    worksheet.merge_cells('J1:Q1')
    worksheet['R1'] = 'CPB Budget'
    worksheet.merge_cells('R1:V1')
    worksheet['C2'] = 'LavieBio'
    worksheet['D2'] = 'Biomica'
    worksheet['E2'] = 'Canonic'
    worksheet['F2'] = 'Agplenus'
    worksheet['G2'] = 'Casterra'
    worksheet['H2'] = 'Agseed'
    worksheet['I2'] = 'Total'
    worksheet['J2'] = 'CSO'
    worksheet['K2'] = 'CSO Microbials'
    worksheet['L2'] = 'MB Upkeep'
    worksheet['M2'] = 'Chempass'
    worksheet['N2'] = 'CP Upkeep'
    worksheet['O2'] = 'Generator'
    worksheet['P2'] = 'GR Upkeep'
    worksheet['Q2'] = 'Total'
    worksheet['R2'] = 'CPBC'
    worksheet['S2'] = 'CPBC Upkeep'
    worksheet['T2'] = 'CPBE'
    worksheet['U2'] = 'CPBE Upkeep'
    worksheet['V2'] = 'Total'
    worksheet['W2'] = 'TOTALS'
    worksheet['A3'] = 'Algorithms Development'
    worksheet.merge_cells('A3:A6')
    worksheet['A7'] = 'Software Development'
    worksheet.merge_cells('A7:A10')
    worksheet['A11'] = 'Bioinformatics'
    worksheet.merge_cells('A11:A14')
    worksheet['B3'] = 'Total Expected (Days)'
    worksheet['B4'] = 'Total Requested (Days)'
    worksheet['B5'] = 'Total Received (Days)'
    worksheet['B6'] = 'Total Postponed (Days) (issues)'
    worksheet['B7'] = 'Total Expected (Days)'
    worksheet['B8'] = 'Total Requested (Days)'
    worksheet['B9'] = 'Total Received (Days)'
    worksheet['B10'] = 'Total Postponed (Days) (issues)'
    worksheet['B11'] = 'Total Expected (Days)'
    worksheet['B12'] = 'Total Requested (Days)'
    worksheet['B13'] = 'Total Received (Days)'
    worksheet['B14'] = 'Total Postponed (Days) (issues)'
    wb.save(path_to_data_from_jira)


# creates Excel funcs in order to calc the sum of days expected and asked in the "sprint planning report" sheet
def days_utilization_report_data_and_calc(path_to_data_from_jira):
    wb = load_workbook(path_to_data_from_jira)
    worksheet = wb['Sprint Planning Report']
    for num in range(3, 15):
        worksheet[f'I{num}'].value = f'=SUM(C{num}:H{num})'
        worksheet[f'Q{num}'].value = f'=SUM(J{num}:P{num})'
        worksheet[f'V{num}'].value = f'=SUM(R{num}:U{num})'
        worksheet[f'W{num}'].value = f'=I{num}+Q{num}+V{num}'
    font_style = Font(name='Verdana', size=10)
    a = Side(border_style='medium', color='000404')
    border = Border(top=a, bottom=a, left=a, right=a)
    for row in worksheet:
        for cell in row:
            cell.font = font_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    wb.save(path_to_data_from_jira)


# takes the priorty and add it to the main sheet in order to decide which missions to approve
def priorty_from_excel(path_to_priorty_excel, path_to_data_from_jira):
    workbook_of_priorty = load_workbook(path_to_priorty_excel)
    worksheet_of_priorty = workbook_of_priorty['Priority']
    dict_of_priorty_and_issue_key = {}
    for num in range(2, worksheet_of_priorty.max_row + 1):
        issue_key = worksheet_of_priorty[f'A{num}'].value
        priorty = worksheet_of_priorty[f'C{num}'].value
        dict_of_priorty_and_issue_key[issue_key] = priorty
    workbook_of_priorty.save(path_to_priorty_excel)
    workbook_of_sprint = load_workbook(path_to_data_from_jira)
    worksheet_of_sprint = workbook_of_sprint['Sheet1']
    for num in range(3, worksheet_of_sprint.max_row + 1):
        key = worksheet_of_sprint[f'K{num}'].value
        if key in dict_of_priorty_and_issue_key:
            worksheet_of_sprint[f'B{num}'].value = dict_of_priorty_and_issue_key[key]
        else:
            worksheet_of_sprint[f'B{num}'].value = 999
    workbook_of_sprint.save(path_to_data_from_jira)


# creates a full data of the progress of the sprint
def planning_and_actual_report(name_of_excel, name_of_excel_from_jira):
    wb = load_workbook(name_of_excel)
    worksheet = wb.create_sheet('Sheet_C')
    worksheet.title = 'Sprint Planning and Execution'
    for col in range(1, 17):
        worksheet[get_column_letter(col) + '1'].fill = PatternFill(patternType='solid', fgColor='DAF7A6')
        worksheet[get_column_letter(col) + '2'].fill = PatternFill(patternType='solid', fgColor='DAF7A6')
        worksheet[get_column_letter(col) + '3'].fill = PatternFill(patternType='solid', fgColor='DAF7A6')
    worksheet['A1'] = 'Data'
    worksheet.merge_cells('A1:O1')
    worksheet['A2'] = 'Planned'
    worksheet.merge_cells('A2:F2')
    worksheet['H2'] = 'Actual'
    worksheet.merge_cells('H2:O2')
    worksheet['A3'] = 'Sprint'
    worksheet['B3'] = 'Project Name'
    worksheet['C3'] = 'Issue Key'
    worksheet['D3'] = 'Summery'
    worksheet['E3'] = 'Budget'
    worksheet['F3'] = 'Days Estimation'
    worksheet['H3'] = 'Project Name'
    worksheet['I3'] = 'Issue Key'
    worksheet['J3'] = 'Summery'
    worksheet['K3'] = 'Budget'
    worksheet['L3'] = 'Work Ratio'
    worksheet['M3'] = 'Days Spent'
    worksheet['N3'] = 'Status'
    worksheet['O3'] = 'Type'
    worksheet['P3'] = 'Comments'
    cs = wb.worksheets[0]
    full_data = []
    counter = 3
    for row in range(3, cs.max_row + 1):
        if cs['G' + str(row)].value == 'approved':
            full_data.append(
                [cs['D' + str(row)].value, cs['K' + str(row)].value, cs['L' + str(row)].value, cs['N' + str(row)].value,
                 cs['H' + str(row)].value, counter])
            counter += 1
    for row in full_data:
        row_to_paste = row[5] + 1
        worksheet["B" + str(row_to_paste)].value = row[0]
        worksheet["C" + str(row_to_paste)].value = row[1]
        worksheet["D" + str(row_to_paste)].value = row[2]
        worksheet["E" + str(row_to_paste)].value = row[3]
        worksheet["F" + str(row_to_paste)].value = row[4]
    full_mid_report_data = grab_line_for_mid_moth_report(name_of_excel_from_jira)
    for row in range(4, worksheet.max_row + 1):
        num = row
        for data_line in full_mid_report_data:
            if data_line[1] == worksheet["C" + str(num)].value:
                worksheet["H" + str(num)].value = data_line[0]
                worksheet["I" + str(num)].value = data_line[1]
                worksheet["J" + str(num)].value = data_line[2]
                worksheet["K" + str(num)].value = data_line[3]
                worksheet["L" + str(num)].value = data_line[4]
                worksheet["M" + str(num)].value = data_line[8]
                worksheet["N" + str(num)].value = data_line[5]
                worksheet["O" + str(num)].value = data_line[6]
                worksheet["F" + str(num)].value = (int(data_line[7]) / 3600) / 8
                if data_line[5] == 'Open':
                    worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='FC2C03')
                elif data_line[5] == 'In Progress':
                    worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='ECB918')
                else:
                    worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='75EE15')
                full_mid_report_data.remove(data_line)
        else:
            continue
    devops_data = []
    bug_data = []
    new_task_data = []
    architect_data = []
    while len(full_mid_report_data) != 0:
        for data_line in full_mid_report_data:
            if str(data_line[9]) == 'taln':
                architect_data.append(data_line)
                full_mid_report_data.remove(data_line)
            elif 'DEV' in str(data_line[1]) or (str(data_line[9]) == 'liavs') or (str(data_line[9]) == 'andreyl') or (
                    str(data_line[9]) == 'vladz'):
                devops_data.append(data_line)
                full_mid_report_data.remove(data_line)
            elif data_line[6] == 'Bug':
                bug_data.append(data_line)
                full_mid_report_data.remove(data_line)
            else:
                new_task_data.append(data_line)
                full_mid_report_data.remove(data_line)
    for data_line in new_task_data:
        num = worksheet.max_row + 1
        worksheet["H" + str(num)].value = data_line[0]
        worksheet["I" + str(num)].value = data_line[1]
        worksheet["J" + str(num)].value = data_line[2]
        worksheet["K" + str(num)].value = data_line[3]
        worksheet["L" + str(num)].value = data_line[4]
        worksheet["M" + str(num)].value = data_line[8]
        worksheet["N" + str(num)].value = data_line[5]
        worksheet["O" + str(num)].value = data_line[6]
        if data_line[5] == 'Open':
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='FC2C03')
        elif data_line[5] == 'In Progress':
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='ECB918')
        else:
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='75EE15')
    worksheet['H' + str(worksheet.max_row + 1)] = 'New Bugs'
    worksheet.merge_cells('H' + str(worksheet.max_row) + ':O' + str(worksheet.max_row))
    for data_line in bug_data:
        num = worksheet.max_row + 1
        worksheet["H" + str(num)].value = data_line[0]
        worksheet["I" + str(num)].value = data_line[1]
        worksheet["J" + str(num)].value = data_line[2]
        worksheet["K" + str(num)].value = data_line[3]
        worksheet["L" + str(num)].value = data_line[4]
        worksheet["M" + str(num)].value = data_line[8]
        worksheet["N" + str(num)].value = data_line[5]
        worksheet["O" + str(num)].value = data_line[6]
        if data_line[5] == 'Open':
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='FC2C03')
        elif data_line[5] == 'In Progress':
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='ECB918')
        else:
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='75EE15')
    worksheet['H' + str(worksheet.max_row + 1)] = 'New Dev Task'
    worksheet.merge_cells('H' + str(worksheet.max_row) + ':O' + str(worksheet.max_row))
    for data_line in devops_data:
        num = worksheet.max_row + 1
        worksheet["H" + str(num)].value = data_line[0]
        worksheet["I" + str(num)].value = data_line[1]
        worksheet["J" + str(num)].value = data_line[2]
        worksheet["K" + str(num)].value = data_line[3]
        worksheet["L" + str(num)].value = data_line[4]
        worksheet["M" + str(num)].value = data_line[8]
        worksheet["N" + str(num)].value = data_line[5]
        worksheet["O" + str(num)].value = data_line[6]
        if data_line[5] == 'Open':
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='FC2C03')
        elif data_line[5] == 'In Progress':
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='ECB918')
        else:
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='75EE15')
    worksheet['H' + str(worksheet.max_row + 1)] = 'New Architect Task'
    worksheet.merge_cells('H' + str(worksheet.max_row) + ':P' + str(worksheet.max_row))
    for data_line in architect_data:
        num = worksheet.max_row + 1
        worksheet["H" + str(num)].value = data_line[0]
        worksheet["I" + str(num)].value = data_line[1]
        worksheet["J" + str(num)].value = data_line[2]
        worksheet["K" + str(num)].value = data_line[3]
        worksheet["L" + str(num)].value = data_line[4]
        worksheet["M" + str(num)].value = data_line[8]
        worksheet["N" + str(num)].value = data_line[5]
        worksheet["O" + str(num)].value = data_line[6]
        if data_line[5] == 'Open':
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='FC2C03')
        elif data_line[5] == 'In Progress':
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='ECB918')
        else:
            worksheet['L' + str(num)].fill = PatternFill(patternType='solid', fgColor='75EE15')
    font_style = Font(name='Verdana', size=10)
    a = Side(border_style='medium', color='000404')
    border = Border(top=a, bottom=a, left=a, right=a)
    for row in worksheet:
        for cell in row:
            cell.font = font_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    wb.save(name_of_excel)


# creates nested lists of the data that is needed in order to evaluate the progress of the missions in a given month
# the variables are defined as the name of the lines in the excl if a given name is change - you must update  it here
# !! need to add work team in order to improve the graphs !!
# !! need to add logs data in order to monitor the way people are working !!
def grab_line_for_mid_moth_report(name_of_excel_from_jira):
    workbook = load_workbook(name_of_excel_from_jira)
    worksheet = workbook.worksheets[0]
    df = pd.read_excel(name_of_excel_from_jira)
    full_mid_report_data = []
    issue_key_col_no = df.columns.get_loc('Issue key')
    issue_key = xlsxwriter.utility.xl_col_to_name(issue_key_col_no)
    summary_col_no = df.columns.get_loc('Summary')
    summary = xlsxwriter.utility.xl_col_to_name(summary_col_no)
    budget_col_no = df.columns.get_loc('Custom field (Budget)')
    budget = xlsxwriter.utility.xl_col_to_name(budget_col_no)
    work_ratio_col_no = df.columns.get_loc('Work Ratio')
    work_ratio = xlsxwriter.utility.xl_col_to_name(work_ratio_col_no)
    status_ratio_col_no = df.columns.get_loc('Status')
    status = xlsxwriter.utility.xl_col_to_name(status_ratio_col_no)
    issue_type_col_no = df.columns.get_loc('Issue Type')
    issue_type = xlsxwriter.utility.xl_col_to_name(issue_type_col_no)
    original_estimate_col_no = df.columns.get_loc('Original Estimate')
    original_estimate = xlsxwriter.utility.xl_col_to_name(original_estimate_col_no)
    time_spent_col_no = df.columns.get_loc('Time Spent')
    time_spent = xlsxwriter.utility.xl_col_to_name(time_spent_col_no)
    assignee_col_no = df.columns.get_loc('Assignee')
    assignee = xlsxwriter.utility.xl_col_to_name(assignee_col_no)
    pm_names = ['elanitec', 'inbarp', 'anatol', 'michalsho', 'michalslu', 'hamutal.e', 'larisar', 'erane', 'eladm',
                'taliy', 'dashau']
    for row in range(2, worksheet.max_row + 1):
        full_mid_report_data.append(
            [budget_naming_tool(worksheet[str(budget) + str(row)].value), worksheet[str(issue_key) + str(row)].value,
             worksheet[str(summary) + str(row)].value, worksheet[str(budget) + str(row)].value,
             worksheet[str(work_ratio) + str(row)].value, worksheet[str(status) + str(row)].value,
             worksheet[str(issue_type) + str(row)].value, worksheet[str(original_estimate) + str(row)].value,
             worksheet[str(time_spent) + str(row)].value, worksheet[str(assignee) + str(row)].value])
    for row in full_mid_report_data:
        for name in pm_names:
            if str(row[9]) in name:
                full_mid_report_data.remove(row)
    for row in full_mid_report_data:
        try:
            while int(row[8]) > 100:
                row[8] = int(row[8]) / 3600
                row[8] = int(row[8]) / 8
        except TypeError:
            row[8] = 0
        print(row)
    return full_mid_report_data


if __name__ == '__main__':
    main()
    # planning_and_actual_report('Sprint_November_4.xlsx', 'sprint_november_54.xlsx')
