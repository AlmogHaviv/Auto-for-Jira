#!/usr/bin/python3.9

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
import xlsxwriter


# to be deleted

def main(name):
    create_sheet_for_report(name)
    xlsx_writer(name, basic_name_data(name))
    styling_excel(name)
    days_utilization_report(name)
    monthly_resource_data('4PMO.xlsx', name)
    days_utilization_report_data_and_calc(name)


# When really used - need to have argparse
def create_sheet_for_report(name):
    wb = load_workbook(name)
    wb.create_sheet("sheet1")
    ws = wb["sheet1"]
    ws.append(['Notes', 'Priority', 'Tasks client/Owner', 'Task Budget Type', 'Task Resource Type/ Exec Unit', 'Task ',
               'Status', 'Approved Effort', 'Actual Status', 'Actual Effort', 'Jira ID (Issue key)', 'Task Description',
               'Task Creator/Budget Owner', 'Budget ID'])
    wb.save(name)


# If the symbol number changes in the company - need to be updated
def budget_naming_tool(string):
    if ('22' in string) or ('143' in string) or ('95' in string) or ('144' in string) or ('19' in string):
        return 'LavieBio'
    elif ('57' in string) or ('255' in string):
        return 'Biomica'
    elif ('209' in string) or ('198' in string) or ('197' in string) or ('205' in string) or ('199' in string) or (
            '150' in string):
        return 'Canonic'
    elif ('21' in string) or ('211' in string) or ('210' in string) or ('23' in string):
        return 'Agplenus'
    elif ('64' in string) or ('999' in string) or ('278' in string) or ('157' in string) or ('162' in string) or (
            '24' in string) or ('8' in string) or ('129' in string) or ('128' in string):
        return 'Agseed'
    elif '269' in string:
        return 'CSO'
    elif '265' in string:
        return 'CSO Microbials'
    elif '264' in string:
        return 'Chempass'
    elif ('277' in string) or ('223' in string) or ('266' in string):
        return 'Generator'
    elif '274' in string:
        return 'CP Upkeep'
    elif '272' in string:
        return 'GR Upkeep'
    elif '273' in string:
        return 'MB Upkeep'
    elif '276' in string:
        return 'CPBC'
    elif '271' in string:
        return 'CPBC Upkeep'
    elif '275' in string:
        return 'CPBE Upkeep'
    elif '280' in string:
        return 'Generator'
    else:
        return 'Error'


# list of names for the resource types
def task_resource_tool(name):
    wb = load_workbook(name)
    cs = wb.worksheets[0]
    raw_name_data = []
    df = pd.read_excel(name)
    col_no = df.columns.get_loc('Assignee')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        raw_name_data.append(cell.value)
    raw_name_data.pop(0)
    wb.save(name)
    res = []
    for val in raw_name_data:
        if val is not None:
            res.append(val)
    return res


# converting the list of names to the resource types
def task_resource_list(string):
    if ('robertoo' in string) or ('renanam' in string) or ('anatm' in string) or ('itair' in string):
        return f'Algorithms Development ({string})'
    elif ('gala' in string) or ('markb' in string) or ('michala' in string) or ('iliab' in string) or (
            'jonathans' in string) or ('liorr' in string):
        return f'Bioinformatics ({string})'
    elif ('duduz' in string) or ('noama' in string) or ('nerias' in string) or ('hodayam' in string):
        return f'Software Development ({string})'
    elif 'taln' in string:
        return f'Architect ({string})'
    elif ('andreyl' in string) or ('liavs' in string) or ('vladz' in string):
        return f'Dev/SysOps ({string})'
    elif 'iliaz' in string:
        return f'Director of CPBC ({string})'
    else:
        return f'other - {string}'


def naming_tool(string):
    if string == 'CP Upkeep':
        return 'Chempass'
    elif string == 'CPBC Upkeep':
        return 'CPBC'
    elif string == 'MB Upkeep':
        return 'CSO Microbials'
    elif string == 'CPBE Upkeep':
        return 'CPBC'
    else:
        return string


def basic_name_data(name):
    wb = load_workbook(name)
    cs = wb.worksheets[0]
    raw_name_data = []
    df = pd.read_excel(name)
    col_no = df.columns.get_loc('Custom field (Budget)')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        raw_name_data.append(cell.value)
    raw_name_data.pop(0)
    wb.save(name)
    res = []
    for val in raw_name_data:
        if val is not None:
            res.append(val)
    return res


def task_effort_estimation_tool(name):
    wb = load_workbook(name)
    cs = wb.worksheets[0]
    raw_name_data = []
    df = pd.read_excel(name)
    col_no = df.columns.get_loc('Original Estimate')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        raw_name_data.append(cell.value)
    raw_name_data.pop(0)
    wb.save(name)
    res = []
    for val in raw_name_data:
        try:
            new_val = int(val) / 3600
            new_val1 = new_val / 8
            res.append(new_val1)
        except TypeError:
            res.append(0)
    return res


def jira_id(name):
    wb = load_workbook(name)
    cs = wb.worksheets[0]
    jira_id_data = []
    df = pd.read_excel(name)
    col_no = df.columns.get_loc('Issue key')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        jira_id_data.append(cell.value)
    jira_id_data.pop(0)
    wb.save(name)
    res_jira_id = []
    for val in jira_id_data:
        if val != None:
            res_jira_id.append(val)
    return res_jira_id


def task_description(name):
    wb = load_workbook(name)
    cs = wb.worksheets[0]
    task_description_data = []
    df = pd.read_excel(name)
    col_no = df.columns.get_loc('Summary')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        task_description_data.append(cell.value)
    task_description_data.pop(0)
    wb.save(name)
    res_task_description_data = []
    for val in task_description_data:
        if val != None:
            res_task_description_data.append(val)
    return res_task_description_data


def task_creator(name):
    wb = load_workbook(name)
    cs = wb.worksheets[0]
    task_creator_data = []
    df = pd.read_excel(name)
    col_no = df.columns.get_loc('Creator')
    a = xlsxwriter.utility.xl_col_to_name(col_no)
    column_a = cs[a]
    for cell in column_a:
        task_creator_data.append(cell.value)
    task_creator_data.pop(0)
    wb.save(name)
    res_task_creator_data = []
    for val in task_creator_data:
        if val != None:
            res_task_creator_data.append(val)
    return res_task_creator_data


def monthly_resource_func(name_of_company_resource, days_of_work, letter, worksheet):
    if name_of_company_resource >= days_of_work:
        worksheet['G' + str(letter)].value = 'approved'
        worksheet['H' + str(letter)].value = worksheet['F' + str(letter)].value
        worksheet['G' + str(letter)].fill = PatternFill(patternType='solid', fgColor='75EE15')
    elif (name_of_company_resource <= days_of_work) and (name_of_company_resource >= 0):
        worksheet['G' + str(letter)].value = 'approved'
        worksheet['H' + str(letter)].value = name_of_company_resource
        worksheet['A' + str(letter)].value = "The task will start but will not finish"
        worksheet['G' + str(letter)].fill = PatternFill(patternType='solid', fgColor='ECB918')
    else:
        worksheet['G' + str(letter)].value = 'Postponed'
        worksheet['H' + str(letter)].value = 0
        worksheet['G' + str(letter)].fill = PatternFill(patternType='solid', fgColor='FC2C03')


def monthly_resource_data(excel_data_name, excel_sprint_name):
    workbook = load_workbook(excel_data_name, data_only=True)
    worksheet1 = workbook['Capacity Potential by Owner']
    worksheet3 = workbook['Human Resources Planning']
    workbook_for_report = load_workbook(excel_sprint_name, data_only=True)
    worksheet_for_report = workbook_for_report['Sprint Planning Report']
    lavie_bio_algo = (worksheet1['B3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['C3'] = lavie_bio_algo
    lavie_bio_bioinfo = (worksheet1['B4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    biomica_bio_algo = (worksheet1['C3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['D3'] = biomica_bio_algo
    biomica_bio_bioinfo = (worksheet1['C4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    biomica_bio_softwaredev = (worksheet1['C5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    canonic_bio_bioinfo = (worksheet1['D4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    canonic_bio_algo = (worksheet1['D3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['E3'] = canonic_bio_algo
    canonic_bio_softwaredev = (worksheet1['D5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    agplenus_bio_bioinfo = (worksheet1['E4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    agplenus_bio_algo = (worksheet1['E3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['F3'] = agplenus_bio_algo
    agplenus_bio_softwaredev = (worksheet1['E5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    agseed_bio_algo = (worksheet1['G3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['H3'] = agseed_bio_algo
    agseed_bio_bioinfo = (worksheet1['G4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    agseed_bio_softwaredev = (worksheet1['G5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    csomicrobials_bio_bioinfo = (worksheet1['I4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    csomicrobials_bio_algo = (worksheet1['I3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['K3'] = csomicrobials_bio_algo
    csomicrobials_bio_softwaredev = (worksheet1['I5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    csomicrobials_bio_devsys = (worksheet1['I7'].value / worksheet1['T7'].value) * worksheet3['H26'].value
    mbupkeep_bio_bioinfo = (worksheet1['J4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    mbupkeep_bio_algo = (worksheet1['J3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['L3'] = mbupkeep_bio_algo
    mbupkeep_bio_softwaredev = (worksheet1['J5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    chempass_bio_algo = (worksheet1['K3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['M3'] = chempass_bio_algo
    chempass_bio_softwaredev = (worksheet1['K5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    cpupkeep_bio_bioinfo = (worksheet1['L4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    cpupkeep_bio_algo = (worksheet1['L3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['N3'] = cpupkeep_bio_algo
    cpupkeep_bio_softwaredev = (worksheet1['L5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    cpupkeep_bio_devsys = (worksheet1['L7'].value / worksheet1['T7'].value) * worksheet3['H26'].value
    genertor_bio_bioinfo = (worksheet1['M4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    genertor_bio_algo = (worksheet1['M3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['O3'] = genertor_bio_algo
    genertor_bio_softwaredev = (worksheet1['M5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    genertor_bio_devsys = (worksheet1['M7'].value / worksheet1['T7'].value) * worksheet3['H26'].value
    grupkeep_bio_bioinfo = (worksheet1['N4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    grupkeep_bio_algo = (worksheet1['N3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['P3'] = grupkeep_bio_algo
    grupkeep_bio_softwaredev = (worksheet1['N5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    cpbc_bio_bioinfo = (worksheet1['O4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    cpbc_bio_algo = (worksheet1['O3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['R3'] = cpbc_bio_algo
    cpbc_bio_softwaredev = (worksheet1['O5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    cpbc_bio_devsys = (worksheet1['O7'].value / worksheet1['T7'].value) * worksheet3['H26'].value
    cpbc_bio_architect = (worksheet1['O6'].value / worksheet1['T6'].value) * worksheet3['H22'].value
    cpbcupkeep_bio_bioinfo = (worksheet1['P4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    cpbcupkeep_bio_algo = (worksheet1['P3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['S3'] = cpbcupkeep_bio_algo
    cpbcupkeep_bio_architect = (worksheet1['P6'].value / worksheet1['T6'].value) * worksheet3['H22'].value
    cpbcupkeep_bio_softwaredev = (worksheet1['P5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    cpbcupkeep_bio_devsys = (worksheet1['P7'].value / worksheet1['T7'].value) * worksheet3['H26'].value
    cso_bio_softwaredev = (worksheet1['H5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    cpbe_bio_bioinfo = (worksheet1['Q4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    cpbe_bio_algo = (worksheet1['Q3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['T3'] = cpbe_bio_algo
    cpbe_bio_softwaredev = (worksheet1['Q5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    cpbe_bio_devsys = (worksheet1['Q7'].value / worksheet1['T7'].value) * worksheet3['H26'].value
    cpbeupkeep_bio_bioinfo = (worksheet1['R4'].value / worksheet1['T4'].value) * worksheet3['H15'].value
    cpbeupkeep_bio_algo = (worksheet1['R3'].value / worksheet1['T3'].value) * worksheet3['H9'].value
    worksheet_for_report['U3'] = cpbeupkeep_bio_algo
    cpbeupkeep_bio_softwaredev = (worksheet1['R5'].value / worksheet1['T5'].value) * worksheet3['H20'].value
    cpbeupkeep_bio_devsys = (worksheet1['R7'].value / worksheet1['T7'].value) * worksheet3['H26'].value
    ws2 = workbook_for_report['Sheet1']
    range_of_func = ws2['V3'].value
    for letter in range(3, range_of_func + 3):
        row_data = []
        for cell in ws2[letter]:
            row_data.append(cell.value)
        if row_data[3] == 'LavieBio':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(lavie_bio_bioinfo, row_data[5], letter, ws2)
                lavie_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(lavie_bio_algo, row_data[5], letter, ws2)
                lavie_bio_algo -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'Biomica':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(biomica_bio_bioinfo, row_data[5], letter, ws2)
                biomica_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(biomica_bio_algo, row_data[5], letter, ws2)
                biomica_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(biomica_bio_softwaredev, row_data[5], letter, ws2)
                biomica_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'Canonic':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(canonic_bio_bioinfo, row_data[5], letter, ws2)
                canonic_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(canonic_bio_algo, row_data[5], letter, ws2)
                canonic_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(canonic_bio_softwaredev, row_data[5], letter, ws2)
                canonic_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'Agplenus':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(agplenus_bio_bioinfo, row_data[5], letter, ws2)
                agplenus_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(agplenus_bio_algo, row_data[5], letter, ws2)
                agplenus_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(agplenus_bio_softwaredev, row_data[5], letter, ws2)
                agplenus_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'Agseed':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(agseed_bio_bioinfo, row_data[5], letter, ws2)
                agseed_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(agseed_bio_algo, row_data[5], letter, ws2)
                agseed_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(agseed_bio_softwaredev, row_data[5], letter, ws2)
                agseed_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'CSO Microbials':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(csomicrobials_bio_bioinfo, row_data[5], letter, ws2)
                csomicrobials_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(csomicrobials_bio_algo, row_data[5], letter, ws2)
                csomicrobials_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(csomicrobials_bio_softwaredev, row_data[5], letter, ws2)
                csomicrobials_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(csomicrobials_bio_devsys, row_data[5], letter, ws2)
                csomicrobials_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'MB Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(mbupkeep_bio_bioinfo, row_data[5], letter, ws2)
                mbupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(mbupkeep_bio_algo, row_data[5], letter, ws2)
                mbupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(mbupkeep_bio_softwaredev, row_data[5], letter, ws2)
                mbupkeep_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'Chempass':
            if 'Algorithms Development' in row_data[4]:
                monthly_resource_func(chempass_bio_algo, row_data[5], letter, ws2)
                chempass_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(chempass_bio_softwaredev, row_data[5], letter, ws2)
                chempass_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'CP Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpupkeep_bio_bioinfo, row_data[5], letter, ws2)
                cpupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpupkeep_bio_algo, row_data[5], letter, ws2)
                cpupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpupkeep_bio_softwaredev, row_data[5], letter, ws2)
                cpupkeep_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpupkeep_bio_devsys, row_data[5], letter, ws2)
                cpupkeep_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'Generator':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(genertor_bio_bioinfo, row_data[5], letter, ws2)
                genertor_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(genertor_bio_algo, row_data[5], letter, ws2)
                genertor_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(genertor_bio_softwaredev, row_data[5], letter, ws2)
                genertor_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(genertor_bio_devsys, row_data[5], letter, ws2)
                genertor_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'GR Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(grupkeep_bio_bioinfo, row_data[5], letter, ws2)
                grupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(grupkeep_bio_algo, row_data[5], letter, ws2)
                grupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(grupkeep_bio_softwaredev, row_data[5], letter, ws2)
                grupkeep_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'CPBC':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpbc_bio_bioinfo, row_data[5], letter, ws2)
                cpbc_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpbc_bio_algo, row_data[5], letter, ws2)
                cpbc_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpbc_bio_softwaredev, row_data[5], letter, ws2)
                cpbc_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpbc_bio_devsys, row_data[5], letter, ws2)
                cpbc_bio_devsys -= row_data[5]
            elif 'Architect' in row_data[4]:
                monthly_resource_func(cpbc_bio_architect, row_data[5], letter, ws2)
                cpbc_bio_architect -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'CPBC Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_bioinfo, row_data[5], letter, ws2)
                cpbcupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_algo, row_data[5], letter, ws2)
                cpbcupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_softwaredev, row_data[5], letter, ws2)
                cpbcupkeep_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_devsys, row_data[5], letter, ws2)
                cpbcupkeep_bio_devsys -= row_data[5]
            elif 'Architect' in row_data[4]:
                monthly_resource_func(cpbcupkeep_bio_architect, row_data[5], letter, ws2)
                cpbcupkeep_bio_architect -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'CSO':
            if 'Software Development' in row_data[4]:
                monthly_resource_func(cso_bio_softwaredev, row_data[5], letter, ws2)
                cso_bio_softwaredev -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'CPBE':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpbe_bio_bioinfo, row_data[5], letter, ws2)
                cpbe_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpbe_bio_algo, row_data[5], letter, ws2)
                cpbe_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpbe_bio_softwaredev, row_data[5], letter, ws2)
                cpbe_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpbe_bio_devsys, row_data[5], letter, ws2)
                cpbe_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        elif row_data[3] == 'CPBE Upkeep':
            if 'Bioinformatics' in row_data[4]:
                monthly_resource_func(cpbeupkeep_bio_bioinfo, row_data[5], letter, ws2)
                cpbeupkeep_bio_bioinfo -= row_data[5]
            elif 'Algorithms Development' in row_data[4]:
                monthly_resource_func(cpbeupkeep_bio_algo, row_data[5], letter, ws2)
                cpbeupkeep_bio_algo -= row_data[5]
            elif 'Software Development' in row_data[4]:
                monthly_resource_func(cpbeupkeep_bio_softwaredev, row_data[5], letter, ws2)
                cpbeupkeep_bio_softwaredev -= row_data[5]
            elif 'Dev/SysOps' in row_data[4]:
                monthly_resource_func(cpbeupkeep_bio_devsys, row_data[5], letter, ws2)
                cpbeupkeep_bio_devsys -= row_data[5]
            else:
                ws2['G' + str(letter)].value = 'approved - no evogene resources'
                ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
        else:
            ws2['G' + str(letter)].value = 'approved - no evogene resources'
            ws2['H' + str(letter)].value = ws2['F' + str(letter)].value
    worksheet_for_report['C4'] = worksheet_for_report['C3'].value - lavie_bio_algo
    worksheet_for_report['D4'] = worksheet_for_report['D3'].value - biomica_bio_algo
    worksheet_for_report['E4'] = worksheet_for_report['E3'].value - canonic_bio_algo
    worksheet_for_report['F4'] = worksheet_for_report['F3'].value - agplenus_bio_algo
    worksheet_for_report['H4'] = worksheet_for_report['H3'].value - agseed_bio_algo
    worksheet_for_report['K4'] = worksheet_for_report['K3'].value - csomicrobials_bio_algo
    worksheet_for_report['L4'] = worksheet_for_report['L3'].value - mbupkeep_bio_algo
    worksheet_for_report['M4'] = worksheet_for_report['M3'].value - chempass_bio_algo
    worksheet_for_report['N4'] = worksheet_for_report['N3'].value - cpupkeep_bio_algo
    worksheet_for_report['O4'] = worksheet_for_report['O3'].value - genertor_bio_algo
    worksheet_for_report['P4'] = worksheet_for_report['P3'].value - grupkeep_bio_algo
    worksheet_for_report['R4'] = worksheet_for_report['R3'].value - cpbc_bio_algo
    worksheet_for_report['S4'] = worksheet_for_report['S3'].value - cpbcupkeep_bio_algo
    worksheet_for_report['T4'] = worksheet_for_report['T3'].value - cpbe_bio_algo
    worksheet_for_report['U4'] = worksheet_for_report['U3'].value - cpbeupkeep_bio_algo
    workbook_for_report.save(excel_sprint_name)


def xlsx_writer(name, raw_name_data):
    jira_id_data = jira_id(name)
    task_description_data = task_description(name)
    task_creator_data = task_creator(name)
    task_resource_data = []
    actual_status_data = []
    actual_effort_data = []
    for string in task_resource_tool(name):
        task_resource_data.append(task_resource_list(string))
        actual_status_data.append('Pending/Planned')
        actual_effort_data.append('0')
    task_effort_estimation_data = task_effort_estimation_tool(name)
    wb1 = xlsxwriter.Workbook(name)
    worksheet = wb1.add_worksheet()
    bold = wb1.add_format({'bold': True})
    headers = ['Notes', 'Priority', 'Tasks client/Owner', 'Task Budget Type', 'Task Resource Type/ Exec Unit', 'Task Effort Estimation (Days)',
               'Task Status', 'Approved Effort', 'Actual Status', 'Actual Effort', 'Jira ID (Issue key)', 'Task Description',
               'Task Creator/Budget Owner', 'Budget ID']
    worksheet.write_row(1, 0, headers, bold)
    budget_name_data = []
    for item in raw_name_data:
        budget_name_data.append(budget_naming_tool(item))
    name_data = []
    for name in budget_name_data:
        name_data.append(naming_tool(name))
    worksheet.write_column(2, 13, raw_name_data)
    worksheet.write_column(2, 3, budget_name_data)
    worksheet.write_column(2, 2, name_data)
    worksheet.write_column(2, 10, jira_id_data)
    worksheet.write_column(2, 11, task_description_data)
    worksheet.write_column(2, 12, task_creator_data)
    worksheet.write_column(2, 4, task_resource_data)
    worksheet.write_column(2, 5, task_effort_estimation_data)
    worksheet.write_column(2, 8, actual_status_data)
    worksheet.write_column(2, 9, actual_effort_data)
    worksheet.write('V3', len(jira_id_data))
    wb1.close()


# start to learn API to jira
def styling_excel(name):
    wb = load_workbook(name)
    ws = wb['Sheet1']
    for col in range(1, 15):
        ws[get_column_letter(col) + '1'].fill = PatternFill(patternType='solid', fgColor='38D9D1')
    for col in range(1, 15):
        ws[get_column_letter(col) + '2'].fill = PatternFill(patternType='solid', fgColor='DAF7A6')
    ws['C1'] = 'Planning'
    ws.merge_cells('C1:H1')
    ws['I1'] = 'Status'
    ws.merge_cells('I1:J1')
    ws['K1'] = 'ID'
    ws['L1'] = 'Description'
    ws.merge_cells('L1:N1')
    font_style = Font(name='Verdana', size=10)
    a = Side(border_style='medium', color='000404')
    border = Border(top=a, bottom=a, left=a, right=a)
    for row in ws:
        for cell in row:
            cell.font = font_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    ws['C1'] = 'Planning'
    ws.merge_cells('C1:H1')
    wb.save(name)


def open_sheet(name):
    wb = load_workbook(name)
    cs = wb.worksheets[0]
    raw_name_data = []
    for row in range(1, 100):
        for col in range(15, 16):
            char = get_column_letter(col)
            raw_name_data.append(cs[char + str(row)].value)
    raw_name_data.pop(0)
    print(raw_name_data)


def days_utilization_report(name_of_excel):
    wb = load_workbook(name_of_excel)
    worksheet = wb.create_sheet('Sheet_B')
    worksheet.title = 'Sprint Planning Report'
    for col in range(1, 24):
        worksheet[get_column_letter(col) + '1'].fill = PatternFill(patternType='solid', fgColor='DAF7A6')
        worksheet[get_column_letter(col) + '2'].fill = PatternFill(patternType='solid', fgColor='DAF7A6')
    worksheet['A1'] = 'Resource Type'
    worksheet.merge_cells('A1:A2')
    worksheet['C1'] = 'Exogenic Customers Budget'
    worksheet.merge_cells('C1:I1')
    worksheet['J1'] = 'Product Budget'
    worksheet.merge_cells('J1:Q1')
    worksheet['R1'] = 'CPB Budget'
    worksheet.merge_cells('R1:V1')
    worksheet['C2'] = 'LavieBio'
    worksheet['D2'] = 'Biomica'
    worksheet['E2'] = 'Canonic'
    worksheet['F2'] = 'Agplenus'
    worksheet['G2'] = 'Casterra'
    worksheet['H2'] = 'Agseed'
    worksheet['I2'] = 'Total'
    worksheet['J2'] = 'CSO'
    worksheet['K2'] = 'CSO Microbials'
    worksheet['L2'] = 'MB Upkeep'
    worksheet['M2'] = 'Chempass'
    worksheet['N2'] = 'CP Upkeep'
    worksheet['O2'] = 'Generator'
    worksheet['P2'] = 'GR Upkeep'
    worksheet['Q2'] = 'Total'
    worksheet['R2'] = 'CPBC'
    worksheet['S2'] = 'CPBC Upkeep'
    worksheet['T2'] = 'CPBE'
    worksheet['U2'] = 'CPBE Upkeep'
    worksheet['V2'] = 'Total'
    worksheet['W2'] = 'TOTALS'
    worksheet['A3'] = 'Algorithms Development'
    worksheet.merge_cells('A3:A6')
    worksheet['B3'] = 'Total Expected (Days)'
    worksheet['B4'] = 'Total Requested (Days)'
    worksheet['B5'] = 'Total Received (Days)'
    worksheet['B6'] = 'Total Postponed (Days) (issues)'
    wb.save(name_of_excel)


def days_utilization_report_data_and_calc(name_of_excel):
    wb = load_workbook(name_of_excel)
    worksheet = wb['Sprint Planning Report']
    worksheet['I3'] = '=SUM(C3:H3)'
    worksheet['Q3'] = '=SUM(J3:P3)'
    worksheet['V3'] = '=SUM(R3:U3)'
    worksheet['W3'] = '=I3+Q3+V3'
    worksheet['I4'] = '=SUM(C4:H4)'
    worksheet['Q4'] = '=SUM(J4:P4)'
    worksheet['V4'] = '=SUM(R4:U4)'
    worksheet['W4'] = '=I4+Q4+V4'
    font_style = Font(name='Verdana', size=10)
    a = Side(border_style='medium', color='000404')
    border = Border(top=a, bottom=a, left=a, right=a)
    for row in worksheet:
        for cell in row:
            cell.font = font_style
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    wb.save(name_of_excel)


if __name__ == '__main__':
    main('Sprint_September_2.xlsx')
    # monthly_resource_data('4PMO.xlsx', 'Sprint_May_2.xlsx')
    # days_utilization_report('Sprint_September_2.xlsx')


